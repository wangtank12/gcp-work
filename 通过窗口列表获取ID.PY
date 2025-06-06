import pyautogui
import time
from PIL import Image
import win32gui
import win32con
import win32clipboard
from paddleocr import PaddleOCR
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# 设置pyautogui的安全设置
pyautogui.FAILSAFE = True
pyautogui.PAUSE = 0.5

def activate_window(window_title):
    """激活并最大化指定窗口"""
    try:
        hwnd = win32gui.FindWindow(None, window_title)
        if hwnd:
            win32gui.ShowWindow(hwnd, win32con.SW_MAXIMIZE)
            win32gui.SetForegroundWindow(hwnd)
            time.sleep(1)
            return True
        return False
    except Exception as e:
        print(f"激活窗口时出错: {e}")
        return False

def get_clipboard_text():
    """获取剪贴板内容"""
    try:
        win32clipboard.OpenClipboard()
        text = win32clipboard.GetClipboardData(win32con.CF_TEXT)
        win32clipboard.CloseClipboard()
        return text.decode('gbk')
    except Exception as e:
        print(f"获取剪贴板内容出错: {e}")
        return ""

def capture_and_recognize(x, y, width, height):
    """截图并识别文字"""
    try:
        screenshot = pyautogui.screenshot(region=(x, y, width, height))
        temp_path = "temp_screenshot.png"
        screenshot.save(temp_path)
        
        ocr = PaddleOCR(use_angle_cls=True, lang="ch")
        result = ocr.ocr(temp_path, cls=True)
        
        os.remove(temp_path)
        
        if result and result[0]:
            # 返回识别的文字及其位置信息
            return [(line[1][0], line[0]) for line in result[0]]
        return []
    except Exception as e:
        print(f"截图识别出错: {e}")
        return []

def save_to_excel(data):
    """保存数据到Excel文件"""
    try:
        current_time = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"姓名ID对应表_{current_time}.xlsx"
        
        # 创建新的工作簿和工作表
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "姓名ID对应表"
        
        # 设置表头
        headers = ["序号", "姓名", "ID"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # 写入数据
        for row, (name, id_num) in enumerate(data, 2):
            ws.cell(row=row, column=1, value=row-1)  # 序号
            ws.cell(row=row, column=2, value=name)   # 姓名
            ws.cell(row=row, column=3, value=id_num) # ID
            
            # 设置单元格对齐方式
            for col in range(1, 4):
                ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")
        
        # 调整列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        
        # 保存文件
        wb.save(filename)
        print(f"数据已保存到: {filename}")
    except Exception as e:
        print(f"保存Excel文件时出错: {e}")

def main():
    window_title = "门诊医生工作站系统----医生:廖振兴--------川北医学院附属医院"
    
    if not activate_window(window_title):
        print("无法找到或激活指定窗口")
        return

    name_id_pairs = []
    
    # 首先处理当前可见的所有行
    visible_rows = capture_and_recognize(31, 121, 70, 483)  # 调整高度为 483 (原515)
    
    # 处理可见行
    for name, pos in visible_rows:
        if "就诊" in name:
            print("检测到就诊，程序结束")
            if name_id_pairs:
                save_to_excel(name_id_pairs)
            return
            
        # 计算点击坐标
        relative_x = pos[0][0]  # 文字在截图中的x偏移
        relative_y = pos[0][1]  # 文字在截图中的y偏移
        
        click_x = 31 + relative_x  # 31是截图区域的起始x坐标
        click_y = 121 + relative_y + 10  # 加10像素点击到行的中间
        
        pyautogui.click(click_x, int(click_y))
        time.sleep(1.5)
        
        pyautogui.click(675, 56)
        time.sleep(0.5)
        pyautogui.hotkey('ctrl', 'c')
        time.sleep(0.5)
        
        id_num = get_clipboard_text()
        if id_num:
            name_id_pairs.append((name, id_num))
            print(f"记录: {name} - {id_num}")
    
    # 继续处理后续行，每次滚动一行
    while True:
        # 检查第25行（最后一行）
        last_row = capture_and_recognize(31, 604, 70, 32)
        
        # 检查是否是空行
        if not last_row:
            print("检测到空行，程序结束")
            if name_id_pairs:
                save_to_excel(name_id_pairs)
            return
        
        # 检查是否是就诊
        if any("就诊" in text for text, _ in last_row):
            print("检测到就诊，程序结束")
            if name_id_pairs:
                save_to_excel(name_id_pairs)
            return
        
        # 如果不是就诊，处理这一行
        for name, pos in last_row:
            # 点击最后一行
            click_x = 31 + pos[0][0]
            click_y = 604 + 16  # 最后一行的中间位置
            
            pyautogui.click(click_x, int(click_y))
            time.sleep(1.5)
            
            pyautogui.click(675, 56)
            time.sleep(0.5)
            pyautogui.hotkey('ctrl', 'c')
            time.sleep(0.5)
            
            id_num = get_clipboard_text()
            if id_num:
                name_id_pairs.append((name, id_num))
                print(f"记录: {name} - {id_num}")
        
        # 向下滚动一行
        pyautogui.click(270, 625)
        time.sleep(0.5)
        
        # 在程序结束前（两个return的位置前）添加统计信息
        if "就诊" in name or not last_row:  # 使用"就诊"作为结束条件，更容易匹配
            print("\n统计信息:")
            print(f"共记录了 {len(name_id_pairs)} 个姓名和ID")
            print("=" * 30)
            if name_id_pairs:
                save_to_excel(name_id_pairs)
            return

if __name__ == "__main__":
    print("程序将在5秒后开始运行...")
    print("请确保门诊系统窗口已打开")
    time.sleep(5)
    
    try:
        main()
    except KeyboardInterrupt:
        print("\n程序被用户中断")
    except Exception as e:
        print(f"程序运行出错: {e}")
